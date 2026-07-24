import openpyxl
import json
import os
import glob
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()

def strip_tags(s):
    return re.sub(r'<[^>]+>', '', s)

def extract_flat(d, prefix=''):
    items = []
    for k, v in d.items():
        new_key = f'{prefix}.{k}' if prefix else k
        if isinstance(v, dict):
            items.extend(extract_flat(v, new_key))
        elif isinstance(v, str):
            items.append((new_key, v))
        elif isinstance(v, list):
            for idx, item in enumerate(v):
                if isinstance(item, str):
                    items.append((f'{new_key}.{idx}', item))
                elif isinstance(item, dict):
                    items.extend(extract_flat(item, f'{new_key}.{idx}'))
    return items

def set_nested_value(d, key_path, value):
    keys = key_path.split('.')
    current = d
    for i, k in enumerate(keys[:-1]):
        next_k = keys[i+1]
        if k.isdigit():
            k_idx = int(k)
            while len(current) <= k_idx:
                current.append({})
            if not isinstance(current[k_idx], (dict, list)):
                current[k_idx] = {}
            current = current[k_idx]
        else:
            if k not in current or not isinstance(current[k], (dict, list)):
                if next_k.isdigit():
                    current[k] = []
                else:
                    current[k] = {}
            current = current[k]
            
    last_key = keys[-1]
    if last_key.isdigit():
        idx = int(last_key)
        while len(current) <= idx:
            current.append(None)
        current[idx] = value
    else:
        current[last_key] = value

# Load Excel translations
wb = openpyxl.load_workbook('translations.xlsx')
excel_rows = []

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    for r in range(2, sheet.max_row + 1):
        key = sheet.cell(row=r, column=1).value
        en = sheet.cell(row=r, column=2).value
        ja = sheet.cell(row=r, column=3).value
        if en and ja:
            n_en = norm(en)
            excel_rows.append({
                'sheet': sheetname,
                'key': str(key) if key else '',
                'en': n_en,
                'ja': str(ja).strip(),
                'tagless_en': strip_tags(n_en)
            })

excel_sheet_key_map = {}
excel_en_map = {}
excel_tagless_map = {}

for r in excel_rows:
    if r['sheet'] and r['key']:
        excel_sheet_key_map[(r['sheet'].lower().strip(), r['key'].strip())] = r['ja']
    if r['en']:
        if r['en'] not in excel_en_map:
            excel_en_map[r['en']] = r['ja']
    if r['tagless_en']:
        if r['tagless_en'] not in excel_tagless_map:
            excel_tagless_map[r['tagless_en']] = r['ja']

en_files = glob.glob('messages/en/*.json')

report = {
    'summary': {},
    'file_details': {}
}

total_keys_all = 0
total_exact_sheet_key = 0
total_exact_en = 0
total_tagless_en = 0
total_existing_ja_fallback = 0
total_missing_in_excel = 0

for fpath in en_files:
    fname = os.path.basename(fpath).replace('.json', '')
    ja_path = os.path.join('messages', 'ja', f'{fname}.json')
    
    with open(fpath, 'r', encoding='utf-8') as f:
        en_data = json.load(f)
        
    existing_ja_flat_map = {}
    if os.path.exists(ja_path):
        with open(ja_path, 'r', encoding='utf-8') as f:
            existing_ja_data = json.load(f)
            for k, v in extract_flat(existing_ja_data):
                existing_ja_flat_map[k] = v

    en_flat = extract_flat(en_data)
    new_ja_dict = {}
    
    file_report = {
        'total': len(en_flat),
        'matched_sheet_key': [],
        'matched_exact_en': [],
        'matched_tagless_en': [],
        'retained_existing_ja': [],
        'missing_all': []
    }
    
    for key, en_val in en_flat:
        total_keys_all += 1
        n_en = norm(en_val)
        n_en_tagless = strip_tags(n_en)
        
        assigned_ja = None
        
        # 1. Match by Sheet + Key
        sk_hit = excel_sheet_key_map.get((fname.lower(), key))
        if sk_hit:
            assigned_ja = sk_hit
            total_exact_sheet_key += 1
            file_report['matched_sheet_key'].append({'key': key, 'en': en_val, 'ja': assigned_ja})
            
        # 2. Match by Exact English text
        if not assigned_ja and n_en in excel_en_map:
            assigned_ja = excel_en_map[n_en]
            total_exact_en += 1
            file_report['matched_exact_en'].append({'key': key, 'en': en_val, 'ja': assigned_ja})
            
        # 3. Match by Tagless English text
        if not assigned_ja and n_en_tagless in excel_tagless_map:
            assigned_ja = excel_tagless_map[n_en_tagless]
            total_tagless_en += 1
            file_report['matched_tagless_en'].append({'key': key, 'en': en_val, 'ja': assigned_ja})

        # 4. Fallback to pre-existing Japanese text in messages/ja/*.json
        if not assigned_ja and key in existing_ja_flat_map:
            assigned_ja = existing_ja_flat_map[key]
            total_existing_ja_fallback += 1
            file_report['retained_existing_ja'].append({'key': key, 'en': en_val, 'ja': assigned_ja})

        # 5. Completely missing
        if not assigned_ja:
            assigned_ja = en_val
            total_missing_in_excel += 1
            file_report['missing_all'].append({'key': key, 'en': en_val})

        set_nested_value(new_ja_dict, key, assigned_ja)

    with open(ja_path, 'w', encoding='utf-8') as f:
        json.dump(new_ja_dict, f, ensure_ascii=False, indent=2)
        f.write('\n')

    report['file_details'][fname] = file_report

total_human_excel = total_exact_sheet_key + total_exact_en + total_tagless_en

report_md = f"""# Japanese Translation Mapping & Audit Report

> [!NOTE]
> This report summarizes the mapping of human Japanese translations from `translations.xlsx` into `messages/ja/*.json`. The English translation files on your website (`messages/en/*.json`) were **100% untouched and preserved**.

## Summary Statistics

- **Total English Keys across Website**: `{total_keys_all}`
- **Mapped from Human Excel**: `{total_human_excel}` keys (**{(total_human_excel / total_keys_all) * 100:.1f}%**)
- **Retained Pre-existing Japanese Keys**: `{total_existing_ja_fallback}` keys
- **Completely Missing / Untranslated Keys**: `{total_missing_in_excel}` keys

### Matching Methodology Breakdown

| Match Strategy | Count | Percentage | Description |
| :--- | :--- | :--- | :--- |
| **Sheet + Key Match** | `{total_exact_sheet_key}` | `{(total_exact_sheet_key / total_keys_all) * 100:.1f}%` | Direct match by sheet name & key in Excel |
| **Exact English Sentence Match** | `{total_exact_en}` | `{(total_exact_en / total_keys_all) * 100:.1f}%` | Exact English text match anywhere in Excel |
| **Tagless English Match** | `{total_tagless_en}` | `{(total_tagless_en / total_keys_all) * 100:.1f}%` | Text match ignoring HTML/JSX tags |
| **Retained Existing JA** | `{total_existing_ja_fallback}` | `{(total_existing_ja_fallback / total_keys_all) * 100:.1f}%` | Kept valid pre-existing translation from `messages/ja/` |
| **Untranslated Fallback** | `{total_missing_in_excel}` | `{(total_missing_in_excel / total_keys_all) * 100:.1f}%` | New text not found in Excel or existing JA |

---

## File-by-File Details

"""

for fname, d in report['file_details'].items():
    excel_m = len(d['matched_sheet_key']) + len(d['matched_exact_en']) + len(d['matched_tagless_en'])
    retained_m = len(d['retained_existing_ja'])
    missing_m = len(d['missing_all'])
    
    report_md += f"### [`messages/ja/{fname}.json`](file:///d:/Projects/jibbv2/messages/ja/{fname}.json)\n"
    report_md += f"- **Total Keys**: `{d['total']}`\n"
    report_md += f"- **Human Excel Translations**: `{excel_m}`\n"
    report_md += f"- **Retained Existing Japanese**: `{retained_m}`\n"
    report_md += f"- **Missing / Untranslated**: `{missing_m}`\n\n"

with open('translation_mapping_report.md', 'w', encoding='utf-8') as f:
    f.write(report_md)

print("Mapping complete! Report written to translation_mapping_report.md")
