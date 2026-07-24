import openpyxl
import json
import os
import glob
import re

def extract_flat(d, prefix=''):
    items = []
    for k, v in d.items():
        new_key = f'{prefix}.{k}' if prefix else k
        if isinstance(v, dict):
            items.extend(extract_flat(v, new_key))
        elif isinstance(v, str):
            items.append((new_key, v))
        elif isinstance(v, list):
            for idx, item in enumerate(v):
                if isinstance(item, str):
                    items.append((f'{new_key}.{idx}', item))
                elif isinstance(item, dict):
                    items.extend(extract_flat(item, f'{new_key}.{idx}'))
    return items

wb = openpyxl.load_workbook('translations.xlsx')
excel_keys = {}
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    for r in range(2, sheet.max_row+1):
        k = sheet.cell(row=r, column=1).value
        en = sheet.cell(row=r, column=2).value
        ja = sheet.cell(row=r, column=3).value
        if en and ja:
            excel_keys[str(en).strip()] = str(ja).strip()

# Check layout.json for resourcesMenu.media
if "Media & Press" in excel_keys:
    print("Found 'Media & Press' in excel:", excel_keys["Media & Press"])
if "Press releases and media resources" in excel_keys:
    print("Found 'Press releases...' in excel:", excel_keys["Press releases and media resources"])
