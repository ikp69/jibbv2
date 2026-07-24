import openpyxl
import json
import os
import glob
import sys
import re
from difflib import SequenceMatcher

sys.stdout.reconfigure(encoding='utf-8')

def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()

def strip_tags(s):
    return re.sub(r'<[^>]+>', '', s)

wb = openpyxl.load_workbook('translations.xlsx')

excel_entries = []
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    for r in range(2, sheet.max_row + 1):
        key = sheet.cell(row=r, column=1).value
        en = sheet.cell(row=r, column=2).value
        ja = sheet.cell(row=r, column=3).value
        if en and ja:
            n_en = norm(en)
            excel_entries.append((sheetname, str(key) if key else '', n_en, ja, strip_tags(n_en)))

en_files = glob.glob('messages/en/*.json')

def flatten_json(d, prefix=''):
    items = []
    for k, v in d.items():
        new_key = f'{prefix}.{k}' if prefix else k
        if isinstance(v, dict):
            items.extend(flatten_json(v, new_key))
        elif isinstance(v, str):
            items.append((new_key, v))
        elif isinstance(v, list):
            for idx, item in enumerate(v):
                if isinstance(item, str):
                    items.append((f'{new_key}.{idx}', item))
                elif isinstance(item, dict):
                    items.extend(flatten_json(item, f'{new_key}.{idx}'))
    return items

stats = {
    'exact_sheet_key': 0,
    'exact_en': 0,
    'tagless_en': 0,
    'fuzzy_high': 0,
    'unmatched': 0
}

samples = []
unmatched_by_file = {}

for fpath in en_files:
    fname = os.path.basename(fpath).replace('.json', '')
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    flat = flatten_json(data)
    unmatched_by_file[fname] = []
    
    for key, en_val in flat:
        n_en = norm(en_val)
        n_en_tagless = strip_tags(n_en)
        
        # 1. Sheet + Key
        sk_match = [e for e in excel_entries if e[0].lower() == fname.lower() and e[1] == key]
        if sk_match:
            stats['exact_sheet_key'] += 1
            continue
            
        # 2. Exact EN
        en_match = [e for e in excel_entries if e[2] == n_en]
        if en_match:
            stats['exact_en'] += 1
            continue

        # 3. Tagless EN match
        tagless_match = [e for e in excel_entries if e[4] == n_en_tagless]
        if tagless_match:
            stats['tagless_en'] += 1
            continue
            
        # 4. Fuzzy match
        best_ratio = 0
        best_e = None
        for e in excel_entries:
            ratio = SequenceMatcher(None, n_en_tagless.lower(), e[4].lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_e = e
                
        if best_ratio >= 0.85:
            stats['fuzzy_high'] += 1
            if len(samples) < 5:
                samples.append((fname, key, n_en, best_e[2], best_e[3], best_ratio))
        else:
            stats['unmatched'] += 1
            unmatched_by_file[fname].append((key, n_en))

print("Matching Breakdown:")
for k, v in stats.items():
    print(f"  {k}: {v}")

total = sum(stats.values())
match_cnt = total - stats['unmatched']
pct = (match_cnt / total) * 100
print(f"Total English keys: {total}")
print(f"Total Matched: {match_cnt} ({pct:.1f}%)")
print(f"Total Missing/Unmatched: {stats['unmatched']}")

print("\nUnmatched count by JSON file:")
for fname, items in unmatched_by_file.items():
    print(f"  {fname}.json: {len(items)} missing")
